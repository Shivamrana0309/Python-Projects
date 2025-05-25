import openpyxl as xl
import os
from openpyxl.styles import PatternFill

color = PatternFill(start_color="FFC6EFCE",end_color="FFC6EFCE",fill_type="solid")

def sort_excel_by_column(filename, sort_column, output_file, descending=True):
    wb = xl.load_workbook(filename)
    ws = wb.active

    data = list(ws.iter_rows(values_only=True))
    header, rows = data[0], data[1:]

    col_index = header.index(sort_column)

    rows.sort(key=lambda x: x[col_index], reverse=descending)

    for i, row in enumerate([header] + rows, start=1):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)
        
        if i <= 41 and i > 1: 
            for j in range(1, len(row) + 1):
                ws.cell(row=i, column=j).fill = color

    wb.save(output_file)

wb = xl.load_workbook("./py/Student_Data_CGPA.xlsx")
sheet = wb['Sheet1']

total_rows = sheet.max_row+1

for row in range(2,total_rows):

    if(sheet.cell(row,2).value == 'F' or sheet.cell(row,3).value == "Day Scholar"):
        sheet.cell(row,8).value = 0
        sheet.cell(row,9).value = 0
        continue

    avg_gpa=(int(sheet.cell(row,6).value) + int(sheet.cell(row,7).value))/2
    sheet.cell(row,8).value = avg_gpa
    attendance = sheet.cell(row,4).value
    score = avg_gpa * 0.7 + attendance * 0.3
    sheet.cell(row,9).value = score

sheet.cell(1,8).value = "Avg_SGPA"
sheet.cell(1,9).value = "Final_Score"

wb.save("temp.xlsx")
sort_excel_by_column("temp.xlsx","Final_Score","hostel_alloted.xlsx",True)

if os.path.exists("temp.xlsx"):
    os.remove("temp.xlsx")
    print("Temporary file deleted successfully!")
else:
    print("Temporary file is not found!")

ws = xl.load_workbook("hostel_alloted.xlsx")
sheet = ws.active

file = open("Hostel_Alloted_Students.txt",'a')
for students in range(2,42):
    file.write("1. ")
    file.write(sheet.cell(students,1).value)
    file.write("\n")
file.close()
ws.save("hostel_alloted.xlsx")

print("Allotted!!")